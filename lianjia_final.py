import json
from multiprocessing import Pool
import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import pymongo
import csv
from lxml import etree
from openpyxl import load_workbook
from location import GetAddressInfo
import socket
class LianJian(object):

    def __init__(self):
        self.GetAddressInfo=GetAddressInfo()
        self.session=requests.session()
        self.infos = []
        self.fp = open('lianjia.csv','a',encoding='utf-8',newline='')
        self.fieldnames = ['链家编号','标题','小区名称','小区纬度','小区经度','邻近地铁站','所属街道', '所属区域'
                           , '总价', '地铁线路', '每平方售价', '建造时间', '房屋户型', '所在楼层'
                           , '建筑面积', '户型结构', '套内面积', '建筑类型' , '房屋朝向', '建筑结构', '装修情况', '梯户比例'
                           , '配备电梯', '挂牌时间', '交易权属', '上次交易', '房屋用途', '房屋年限', '产权所属','url','别墅类型', '抵押信息']
        self.writer = csv.DictWriter(self.fp,fieldnames = self.fieldnames)
        self.writer.writeheader()
        self.pd_look = pd.DataFrame(columns=self.fieldnames)
    def generate_allurl(self,user_in_nub, user_in_city):  # 生成url
        url = 'http://' + user_in_city + '.lianjia.com/ershoufang/pg{}/'
        for url_next in range(80, int(user_in_nub)):
            yield url.format(url_next)
            
    
    
    def get_allurl(self,generate_allurl):  # 分析url解析出每一页的详细url
    
        #这里模拟一下请求头，头文件是从浏览器里面抓到的，否则服务会回复403错误，（其实就是服务器做的简单防爬虫检测）
        headers={
            'Host':'sh.lianjia.com',
            'Connection':'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
            'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cookie':'TY_SESSION_ID=a026b393-5632-4f1a-a8b0-0849922077b2; lianjia_uuid=2519cfc6-da73-4c95-afed-33ec70e1f70c; _smt_uid=5eb51c9e.4a57a330; UM_distinctid=171f377ccbc214-04b2b9efa5eff5-3a614f0b-144000-171f377ccbd16c; _ga=GA1.2.339410402.1588927649; _gid=GA1.2.505688922.1590473956; _jzqx=1.1590485226.1590485226.1.jzqsr=sh%2Elianjia%2Ecom|jzqct=/ershoufang/.-; select_city=310000; _jzqckmp=1; digv_extends=%7B%22utmTrackId%22%3A%2221583074%22%7D; CNZZDATA1253492439=2138692411-1588923379-https%253A%252F%252Fwww.baidu.com%252F%7C1590632998; CNZZDATA1254525948=16081044-1588923973-https%253A%252F%252Fwww.baidu.com%252F%7C1590631332; CNZZDATA1255633284=273650142-1588926847-https%253A%252F%252Fwww.baidu.com%252F%7C1590633764; CNZZDATA1255604082=931148808-1588927291-https%253A%252F%252Fwww.baidu.com%252F%7C1590634139; _qzjc=1; _jzqa=1.3485495383625437000.1590567682.1590567682.1590635170.2; _jzqc=1; _jzqy=1.1588927647.1590635170.5.jzqsr=baidu.jzqsr=baidu|jzqct=%E9%93%BE%E5%AE%B6%E4%BA%8C%E6%89%8B%E6%88%BF; Hm_lvt_9152f8221cb6243a53c83b956842be8a=1590482198,1590540549,1590567686,1590635172; lianjia_ssid=0194ef77-7bf9-4348-8782-49f901ceca0a; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%22171f377ce7310a-0266c8a467357b-3a614f0b-1327104-171f377ce76524%22%2C%22%24device_id%22%3A%22171f377ce7310a-0266c8a467357b-3a614f0b-1327104-171f377ce76524%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E4%BB%98%E8%B4%B9%E5%B9%BF%E5%91%8A%E6%B5%81%E9%87%8F%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fsp0.baidu.com%2F9q9JcDHa2gU2pMbgoY3K%2Fadrc.php%3Ft%3D06KL00c00fZE91Y02nIM0nVfAsakQnqX00000FLdx7C00000V22azh.THLKVQ1iVEe2doXO0A3qrj6srHnzPH7xPj-xpA7EgLKM0ZnqmvP-PhwhuHfsnj0kPHFbnfKd5Hnzf1KanWDLfWn4P1c%22%2C%22%24latest_referrer_host%22%3A%22sp0.baidu.com%22%2C%22%24latest_search_keyword%22%3A%22%E9%93%BE%E5%AE%B6%E4%BA%8C%E6%89%8B%E6%88%BF%22%2C%22%24latest_utm_source%22%3A%22baidu%22%2C%22%24latest_utm_medium%22%3A%22pinzhuan%22%2C%22%24latest_utm_campaign%22%3A%22wyshanghai%22%2C%22%24latest_utm_content%22%3A%22biaotimiaoshu%22%2C%22%24latest_utm_term%22%3A%22biaoti%22%7D%7D; _gat=1; _gat_past=1; _gat_global=1; _gat_new_global=1; _gat_dianpu_agent=1; Hm_lpvt_9152f8221cb6243a53c83b956842be8a=1590635705; _qzja=1.73996101.1590567681410.1590567681411.1590635169645.1590635702058.1590635704791.0.0.0.8.2; _qzjb=1.1590635169645.4.0.0.0; _qzjto=4.1.0; _jzqb=1.4.10.1590635170.1; srcid=eyJ0Ijoie1wiZGF0YVwiOlwiZjg4YTcyYzhjYmU5NzNjMTM0YTZlOTY1YTBkNzVhMDVhMDI1OWM0MDgxN2VlYWQzZTZjZmZhZGRjNWNjNzUyZjBmYzkzMzRkYTU4NGNkYzI3OTI3NWRkY2U1ZTQ5OTU3YTY1YjI4MmU5ODczNTY1NGZiNmI4MTg4YjkzODdlOWZmN2ZlM2YxNjdkZjA0OTBhZjI2YmFmMGIwM2QxOTBmMTFiZjliODhiYTJmYmVkMTZiZTJhNWE4NDY4NjAyMDEyZjI1M2YwYjZiYzNiMGNjOTdmMWJhMzI3NWQ0YTA1ZTc4NWE0ZDU3NTI4YTliYzk2Mjc5M2IzZGM3OTkxMWE0M2VlYTAwYmMwMWI3ZjA2MjBjMTFjODI4NzU2OTIyNTQ3ZTNjOTExZGI5MWFkNmNmNTFlZjdjM2IzYTk0YmI0NDI3NDZjNGY0ODliZDI1ODQwMmRiNWRiYjJiMzE4ZDM5ZFwiLFwia2V5X2lkXCI6XCIxXCIsXCJzaWduXCI6XCJkMDQ2NTE4OVwifSIsInIiOiJodHRwczovL3NoLmxpYW5qaWEuY29tL2Vyc2hvdWZhbmcvIiwib3MiOiJ3ZWIiLCJ2IjoiMC4xIn0='
        }
        self.session.headers.clear()
        self.session.headers.update(headers)
        get_url = self.session.get(generate_allurl)
        if get_url.status_code == 200:
            soup = BeautifulSoup(get_url.text, 'html5lib')
            urls=[]
            liitems = soup.select('li.clear.LOGCLICKDATA')
            for item in liitems:
                url = item.select('a')[0]['href']
                urls.append(url)
            return urls

    def open_url(self,re_get):  # 分析详细url获取所需信息
        socket.setdefaulttimeout(30)
        res = self.session.get(re_get,timeout = 30)
        info = {}
        soup = BeautifulSoup(res.text, 'lxml')
        text = res.content.decode('UTF-8')
        tree = etree.HTML(text)
        try:
            info['链家编号'] = str(re_get)[34:].rsplit('.html')[0]
            info['标题'] = soup.select('.main')[0].text
            info['总价'] = soup.select('.total')[0].text + '万'
            info['每平方售价'] = soup.select('.unitPriceValue')[0].text
            info['每平方售价'] = re.search(r'\d+',info['每平方售价'])[0]
            try:
                info['建造时间'] = soup.select('.subInfo')[2].text
            except:
                info['建造时间'] = re.search(r'\d+',info['建造时间'])[0]
            info['小区名称'] = soup.select('.info')[0].text
            info['小区纬度'],info['小区经度']=self.GetAddressInfo.run(name=info['小区名称'])#调用类函数GetAddressInfo获得小区经纬度和地址
            loc = soup.select('.info a')[0].text + ':' + soup.select('.info a')[1].text
            info['所属区域'] = re.split(':',loc)[0]
            info['所属街道'] = re.split(':',loc)[1]
            try:
                traffic = tree.xpath("//a[@class='supplement']/text()")[0]
                info['地铁线路'] = re.search(r'\d+',traffic)[0]
                info['邻近地铁站'] = re.search(r'[线\)]\w+',traffic)[0][1:]
            except:
                info['地铁线路'] = '暂无信息'
                info['邻近地铁站'] = '暂无信息'
            base_items= tree.xpath("//div[@class='content']//ul/li//text()")
            base_items =list(filter(lambda x: re.match(r"\w", x), base_items)) 
            try:
                for index,item in enumerate(base_items):
                    item = item.strip()
                    index=2*index
                    if index == 34:
                        break
                    else:
                        info[base_items[index]] = base_items[index+1]  
            except:
                print('出错:',re_get)
            info['url'] = re_get
            self.writer.writerow(info)
            print(info)
        except:
            pass
        return info

    def info_to_csv(self,info):
        if len(self.infos) >= 100:
            self.writer.writerows(self.infos)
            self.infos.clear()
        self.infos.append(info)

    def append_df_to_excel(self, info, headers=None,sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, 
                        **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]
        Returns: None
        """
    # from openpyxl import load_workbook
    
    # import pandas as pd
    
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
        filename = r'D:\vscode\GitHub\链家二手房.xlsx'
        pd_look=self.pd_look.append(info,ignore_index=True)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError
    
    
        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)
    
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
    
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
    
            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass
    
        if startrow is None:
            startrow = 0
    
        # write out the new sheet
        pd_look.to_excel(writer, sheet_name, header=headers,startrow=startrow, **to_excel_kwargs)
    
        # save the workbook
        writer.save()

    def writer_to_text(self,list):  # 储存到text
        with open('链家二手房.text', 'a', encoding='utf-8')as f:
            f.write(json.dumps(list, ensure_ascii=False) + '\n')
            f.close()
    
    def run(self):
        count = 0
        num = 0
        fail_urls = []
        for i in self.generate_allurl('100', 'sh'):
            count = count+1
            print('正在爬取第{0}页'.format(count))
            print('='*50)
            for index,url in enumerate(self.get_allurl(i)):
                
                index = index +1
                print('正在爬取第{0}页的第{1}条'.format(count,index))
                try:
                    info = self.open_url(url)
                    self.writer_to_text(info)    #储存到text文件
                    self.info_to_csv(info)
                    if index==1 and count ==1:
                        #excel第一行插入特征名称
                        self.append_df_to_excel(info, headers=self.fieldnames, index = False)
                    else:
                        self.append_df_to_excel(info, index = False,headers=None)
                except:
                    fail_urls = fail_urls.append(url)
                    num = num+1
                    print('爬取第{0}页的第{1}条失败，已经略过,目前已略过{2}条'.format(count,index,num),url)
            print("第{}页已经爬取完成".format(count)) 
        self.info_to_csv(fail_urls)
if __name__ == '__main__':
    lianjia = LianJian()
    lianjia.run()
        
